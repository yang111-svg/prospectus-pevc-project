#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
JSONL 转 Excel 脚本
===================
读取 outputs/week2_jsonl/ 目录下所有 .jsonl 文件，
按公司分组生成 Excel 文件（认缴流量 / 股权结构存量 / Cross-check 三表）。

用法:
    python scripts/jsonl_to_excel.py
    python scripts/jsonl_to_excel.py --company 603418
    python scripts/jsonl_to_excel.py --input-dir outputs/week2_jsonl --output-dir outputs/week2_excel
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 项目根目录
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).resolve().parent.parent  # prospectus-pevc-project

# ---------------------------------------------------------------------------
# 默认路径
# ---------------------------------------------------------------------------
DEFAULT_INPUT_DIR = BASE_DIR / "outputs" / "week2_jsonl"
DEFAULT_OUTPUT_DIR = BASE_DIR / "outputs" / "week2_excel"

# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=11, name="微软雅黑")
HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")  # 浅蓝底色
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CELL_ALIGNMENT_RIGHT = Alignment(horizontal="right", vertical="center")
CELL_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Sheet 列定义
# ---------------------------------------------------------------------------

# 认缴流量 sheet 列顺序
SUBSCRIPTION_FLOW_COLUMNS = [
    ("pdf_page", "PDF页码"),
    ("event_date", "增资日期"),
    ("subscriber_name", "认购方"),
    ("subscriber_type", "认购方类型"),
    ("subscription_shares", "认购数量(万股)"),
    ("subscription_amount", "认购金额(万元)"),
    ("subscription_price", "认购价格(元/股)"),
    ("evidence_text", "原文证据"),
    ("notes", "备注"),
]

# 股权结构存量 sheet 列顺序
EQUITY_SNAPSHOT_COLUMNS = [
    ("pdf_page", "PDF页码"),
    ("snapshot_label", "时点"),
    ("snapshot_date", "时点日期"),
    ("basis_desc", "口径描述"),
    ("total_shares", "总股本(万股)"),
    ("total_capital", "总出资额(万元)"),
    ("shareholder_name", "股东名称"),
    ("shareholder_type", "股东类型"),
    ("holding_shares", "持股数(万股)"),
    ("holding_capital", "出资额(万元)"),
    ("holding_ratio", "持股比例"),
    ("evidence_text", "原文证据"),
    ("notes", "备注"),
]

# Cross-check sheet 列顺序
CROSS_CHECK_COLUMNS = [
    ("check_type", "检查类型"),
    ("before_snapshot", "变更前时点"),
    ("after_snapshot", "变更后时点"),
    ("field_name", "字段名"),
    ("expected_value", "预期值"),
    ("actual_value", "实际值"),
    ("difference", "差额"),
    ("status", "状态"),
]

# ---------------------------------------------------------------------------
# 需要特殊格式化的列（百分比 / 金额）
# ---------------------------------------------------------------------------
PERCENTAGE_COLUMNS = {"holding_ratio"}  # 持股比例 -> 百分比格式
AMOUNT_COLUMNS = {
    "subscription_shares", "subscription_amount", "subscription_price",
    "total_shares", "total_capital",
    "holding_shares", "holding_capital",
    "difference",
}  # 金额/数量 -> 保留2位小数

# 数值列右对齐
NUMERIC_COLUMNS = PERCENTAGE_COLUMNS | AMOUNT_COLUMNS


# ---------------------------------------------------------------------------
# 核心函数
# ---------------------------------------------------------------------------

def read_jsonl_files(input_dir: Path) -> dict:
    """
    读取 input_dir 下所有 .jsonl 文件，按公司分组返回。

    返回:
        dict: { "股票代码_公司简称": [record1, record2, ...] }
        每条 record 是一个 dict，至少包含 "record_type" 字段。
    """
    if not input_dir.exists():
        print(f"[错误] 输入目录不存在: {input_dir}")
        sys.exit(1)

    jsonl_files = sorted(input_dir.glob("*.jsonl"))
    if not jsonl_files:
        print(f"[警告] 输入目录下没有 .jsonl 文件: {input_dir}")
        return {}

    company_data = defaultdict(list)

    for jsonl_file in jsonl_files:
        print(f"  读取: {jsonl_file.name}")
        with open(jsonl_file, "r", encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                try:
                    record = json.loads(line)
                except json.JSONDecodeError as e:
                    print(f"    [警告] {jsonl_file.name} 第{line_num}行 JSON 解析失败: {e}")
                    continue

                # 提取公司标识
                stock_code = record.get("stock_code", "")
                company_short_name = record.get("company_short_name", "")
                company_key = f"{stock_code}_{company_short_name}" if stock_code else jsonl_file.stem

                record["_source_file"] = jsonl_file.name
                company_data[company_key].append(record)

    return dict(company_data)


def classify_records(records: list) -> dict:
    """
    将记录按 record_type 分类。

    返回:
        dict: {
            "subscription_flow": [...],
            "equity_snapshot": [...],
            "cross_check": [...],
            "other": [...]
        }
    """
    classified = {
        "subscription_flow": [],
        "equity_snapshot": [],
        "cross_check": [],
        "other": [],
    }

    for record in records:
        record_type = record.get("record_type", "")
        if record_type == "subscription_flow":
            classified["subscription_flow"].append(record)
        elif record_type == "equity_snapshot":
            classified["equity_snapshot"].append(record)
        elif record_type == "cross_check":
            classified["cross_check"].append(record)
        else:
            classified["other"].append(record)

    return classified


def write_sheet(ws, columns_def: list, rows_data: list) -> None:
    """
    向 worksheet 写入数据并应用格式。

    参数:
        ws: openpyxl Worksheet 对象
        columns_def: 列定义列表，每项为 (字段名, 显示名)
        rows_data: 数据行列表，每项为 dict
    """
    field_names = [col[0] for col in columns_def]
    display_names = [col[1] for col in columns_def]

    # ---- 写表头 ----
    for col_idx, header in enumerate(display_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # ---- 写数据行 ----
    for row_idx, row_data in enumerate(rows_data, 2):
        for col_idx, field_name in enumerate(field_names, 1):
            value = row_data.get(field_name, "")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            # 格式化
            if field_name in PERCENTAGE_COLUMNS:
                # 持股比例：如果值是数字且在 0~1 之间，视为小数比例；否则视为百分比值
                if isinstance(value, (int, float)):
                    if 0 < value < 1:
                        cell.number_format = '0.00%'
                    else:
                        # 值已经是百分比形式（如 15.00 表示 15%），转为小数再格式化
                        cell.value = value / 100.0 if value else 0
                        cell.number_format = '0.00%'
                    cell.alignment = CELL_ALIGNMENT_RIGHT
                else:
                    cell.alignment = CELL_ALIGNMENT_LEFT

            elif field_name in AMOUNT_COLUMNS:
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    cell.alignment = CELL_ALIGNMENT_RIGHT
                else:
                    cell.alignment = CELL_ALIGNMENT_LEFT

            elif field_name in NUMERIC_COLUMNS:
                cell.alignment = CELL_ALIGNMENT_RIGHT

            else:
                cell.alignment = CELL_ALIGNMENT_LEFT

    # ---- 列宽自适应 ----
    for col_idx, field_name in enumerate(field_names, 1):
        # 计算最大宽度
        max_len = len(str(display_names[col_idx - 1]))  # 表头宽度
        for row_data in rows_data:
            val = row_data.get(field_name, "")
            val_str = str(val) if val is not None else ""
            # 中文字符算2个宽度单位
            char_len = sum(2 if ord(c) > 127 else 1 for c in val_str)
            max_len = max(max_len, char_len)

        # 设置列宽（加一些 padding）
        adjusted_width = min(max_len + 4, 60)  # 最大60字符宽
        adjusted_width = max(adjusted_width, 10)  # 最小10字符宽
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # ---- 冻结首行 ----
    ws.freeze_panes = "A2"


def generate_excel(company_key: str, records: list, output_dir: Path) -> Path:
    """
    为单家公司生成 Excel 文件。

    参数:
        company_key: 公司标识，格式 "股票代码_公司简称"
        records: 该公司的所有 JSONL 记录列表
        output_dir: 输出目录

    返回:
        Path: 生成的 Excel 文件路径
    """
    wb = Workbook()

    # ---- 分类记录 ----
    classified = classify_records(records)

    # ---- Sheet1: 认缴流量 ----
    ws1 = wb.active
    ws1.title = "认缴流量"
    write_sheet(ws1, SUBSCRIPTION_FLOW_COLUMNS, classified["subscription_flow"])

    # ---- Sheet2: 股权结构存量 ----
    ws2 = wb.create_sheet(title="股权结构存量")
    write_sheet(ws2, EQUITY_SNAPSHOT_COLUMNS, classified["equity_snapshot"])

    # ---- Sheet3: Cross-check ----
    ws3 = wb.create_sheet(title="Cross-check")
    write_sheet(ws3, CROSS_CHECK_COLUMNS, classified["cross_check"])

    # ---- 如果有 other 类型记录，额外创建一个 sheet ----
    if classified["other"]:
        ws_other = wb.create_sheet(title="其他记录")
        # 动态生成列定义
        all_keys = set()
        for rec in classified["other"]:
            all_keys.update(k for k in rec.keys() if not k.startswith("_"))
        other_columns = [(k, k) for k in sorted(all_keys)]
        write_sheet(ws_other, other_columns, classified["other"])

    # ---- 保存文件 ----
    output_file = output_dir / f"{company_key}_三表.xlsx"
    wb.save(output_file)
    return output_file


def main():
    parser = argparse.ArgumentParser(
        description="JSONL 转 Excel 脚本：将 week2_jsonl 目录下的 JSONL 文件按公司分组生成三表 Excel",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python scripts/jsonl_to_excel.py
  python scripts/jsonl_to_excel.py --company 603418
  python scripts/jsonl_to_excel.py --input-dir outputs/week2_jsonl --output-dir outputs/week2_excel
        """,
    )
    parser.add_argument(
        "--input-dir",
        type=str,
        default=str(DEFAULT_INPUT_DIR),
        help=f"JSONL 输入目录（默认: {DEFAULT_INPUT_DIR}）",
    )
    parser.add_argument(
        "--output-dir",
        type=str,
        default=str(DEFAULT_OUTPUT_DIR),
        help=f"Excel 输出目录（默认: {DEFAULT_OUTPUT_DIR}）",
    )
    parser.add_argument(
        "--company",
        type=str,
        default=None,
        help="只转换指定公司（股票代码，如 603418）",
    )

    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    # 确保输出目录存在
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("JSONL 转 Excel 工具")
    print("=" * 60)
    print(f"输入目录: {input_dir}")
    print(f"输出目录: {output_dir}")
    print()

    # ---- 读取所有 JSONL 文件 ----
    print("[1/3] 读取 JSONL 文件...")
    company_data = read_jsonl_files(input_dir)
    if not company_data:
        print("没有找到任何数据，退出。")
        sys.exit(0)

    print(f"  共 {len(company_data)} 家公司\n")

    # ---- 筛选指定公司 ----
    if args.company:
        matched_keys = [k for k in company_data if k.startswith(f"{args.company}_")]
        if not matched_keys:
            # 尝试模糊匹配
            matched_keys = [k for k in company_data if args.company in k]
        if not matched_keys:
            print(f"[错误] 未找到股票代码包含 '{args.company}' 的公司")
            print(f"  可用公司: {', '.join(company_data.keys())}")
            sys.exit(1)
        company_data = {k: company_data[k] for k in matched_keys}
        print(f"  筛选公司: {', '.join(matched_keys)}\n")

    # ---- 生成 Excel ----
    print("[2/3] 生成 Excel 文件...")
    success_count = 0
    for company_key, records in company_data.items():
        try:
            output_file = generate_excel(company_key, records, output_dir)

            # 统计各类记录数
            classified = classify_records(records)
            n_sub = len(classified["subscription_flow"])
            n_eq = len(classified["equity_snapshot"])
            n_cc = len(classified["cross_check"])
            n_other = len(classified["other"])

            print(f"  {company_key}: 认缴流量={n_sub}, 股权结构存量={n_eq}, "
                  f"Cross-check={n_cc}, 其他={n_other} -> {output_file.name}")
            success_count += 1
        except Exception as e:
            print(f"  [错误] {company_key}: {e}")

    # ---- 汇总 ----
    print()
    print("[3/3] 完成")
    print(f"  成功: {success_count}/{len(company_data)} 家公司")
    print(f"  输出目录: {output_dir}")


if __name__ == "__main__":
    main()
